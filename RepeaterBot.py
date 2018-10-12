import httplib2
import sys
import os
import hashlib
import string
from urllib import request
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors, Color

http = httplib2.Http()

_MIN_COL_WIDTH = 30
_LETTERS = string.ascii_uppercase
_KEYWORDS = [('Downlink', 'downlink'), 
        ('Uplink', 'uplink'),
        ('Offset', 'offset'),
        ('Uplink Tone', 'uplinktone'),
        ('Downlink Tone', 'downlinktone'),
        ('Call', 'call'),
        ('Use', 'use'),
        ('Sponsor', 'sponsor'),
        ('Affaliate', 'affaliate'),
        ('Links', 'links'),
        ('EchoLink', 'echolink'),
        ('IRPL', 'irpl'),
        ('Last update','lastupdate')]
_FONT_HEADER = Font(name='Calibri',
        color=colors.BLUE,
        size=18)
_FONT_BODY = Font(name='Calibri',
        size=14)
def _AS_TEXT(value): return str(value) if value is not None else ""

class RepeaterBot:
    rb_url = 'https://www.repeaterbook.com'
    search_results = []

    def __init__(self, callsigns=None):
        self.callsigns = callsigns

        # search repeaterbook for all callsigns
        self.search_results = [(sign, self.search_rb(sign)) for sign in self.callsigns]

        row = 2
        self.repeaters = []
        for result in self.search_results:
            self.repeaters += [self.load_rb(url) for url in result[1]]

    def _read_cache(self, key):
        cache = open(".cache/%s" % hashlib.md5(key.encode()).hexdigest(), 'r')
        res = cache.read()
        cache.close()

        return res

    def _write_cache(self, key, val):
        if not os.path.exists('.cache'):
                os.makedirs('.cache')

        cache = open(".cache/%s" % hashlib.md5(key.encode()).hexdigest(), 'w')
        cache.write(str(val))
        cache.close()

    def _request(self, url):
        # check if request is cached
        try:
            status = None
            response = self._read_cache(url)
        except FileNotFoundError:
            # ok lets just make a new one then
            status, response = http.request(url)
            self._write_cache(url, response)
        
        return (status, response)

    def search_rb(self, callsign):
        res = self._request("%s/repeaters/callResult.php?call=%s&submit=RepeaterBook" % (self.rb_url, callsign))
        soup = BeautifulSoup(res[1], 'html.parser')

        links = []

        for x in soup.find_all('a', title='View details'):
            links.append(x['href'])

        return links

    def load_rb(self, rb_href):
        res = self._request("%s/repeaters/%s" % (self.rb_url, rb_href))
        soup = BeautifulSoup(res[1], 'html.parser')

        details = {}
        for keyword in _KEYWORDS:
            details[keyword[1]] = None

        for x in soup.find_all('td'):
            rb_cell = str(x.encode_contents().strip())

            for keyword in _KEYWORDS:
                cell_title = "%s:" % keyword[0]
                if rb_cell.find(cell_title) != -1 and x.find_next_sibling():
                    details[keyword[1]] = x.find_next_sibling().renderContents().strip()
                    print(x.find_next_sibling().find('a'))

        return details

    def print_repeaters(self):
        for repeater in self.repeaters:
            for word in _KEYWORDS:
                print("%s: %s" % (word[0], repeater[word[1]]))

            print("---------")

    def save_excel(self, wb_dest='repeaters.xlsx'):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = 'Repeaters'

        # build headers
        for word in _KEYWORDS:
            coords = "%s1" % _LETTERS[_KEYWORDS.index(word)]
            self.ws[coords].font = _FONT_HEADER
            self.ws[coords] = word[0]

        row = 2
        for repeater in self.repeaters:
            for word in _KEYWORDS:
                col = _LETTERS[_KEYWORDS.index(word)]
                coords = "%s%d" % (col, row)

                self.ws[coords].font = _FONT_BODY
                self.ws[coords] = repeater[word[1]]

            row += 1

        for column_cells in self.ws.iter_cols(max_col=len(_KEYWORDS), max_row=row):
                length = max(len(_AS_TEXT(cell.value)) for cell in column_cells) * 1.5

                if length > _MIN_COL_WIDTH:
                    self.ws.column_dimensions[column_cells[0].column].width = length
                else:
                    self.ws.column_dimensions[column_cells[0].column].width = _MIN_COL_WIDTH

        self.wb.save(filename = wb_dest)


def main():
    # load callsigns from input file
    with open(sys.argv[1]) as f:
        callsigns = f.readlines()

    # start repeaterbot and pass callsigns
    rpb = RepeaterBot([x.strip() for x in callsigns])

    # print to screen and excel file
    rpb.print_repeaters()
    rpb.save_excel()

if __name__ == "__main__" :
    main()
